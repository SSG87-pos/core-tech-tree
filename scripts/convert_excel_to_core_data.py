#!/usr/bin/env python3
"""Convert the one-sheet PTRS Core Tech Excel template into HTML-ready JSON.

Usage:
  python3 scripts/convert_excel_to_core_data.py \
    outputs/core-tech-tree-template/PTRS_Core_Tech_Data_Template.xlsx
"""

from __future__ import annotations

import argparse
import json
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


DEFAULT_OUTPUT = Path("data/core-tech-data.json")
DEFAULT_RESEARCH_GROUPS = [
    "수소환원제철",
    "차세대철강",
    "제선",
    "제강",
    "엔지니어링",
    "로봇AI",
    "열연선재후판",
    "냉연도금",
    "STS연구",
    "자동차소재",
    "표면",
    "자동차소재솔루션",
    "강재솔루션",
    "강건재솔루션",
]
DEFAULT_TEAMS = [
    "[C]용선온도하락방지",
    "[C]Inocast",
    "[C]수소환원",
    "[C]ESF연구",
    "[C]미래연원료",
    "[8대]에너지용후판",
    "[8대]고Mn강",
    "[C]K방산제품",
    "[8대]신재생에너지PosMAC",
    "[C]신도금",
    "[8대]HyperNO",
    "[8대]전력용전기강판",
    "[C]원자력재료",
    "[8대]차세대성장시장용STS",
    "[8대]GigaSteel",
    "[8대]전기로고급강",
    "[C]강구조아파트",
    "[C]제품AX",
]


def clean(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def as_bool(value: Any) -> bool:
    return clean(value).lower() in {"true", "y", "yes", "1", "core"}


def norm_optional(value: Any) -> str:
    text = clean(value)
    return text if text and text != "-" else "미지정"


def table_rows(sheet) -> list[dict[str, Any]]:
    headers = [clean(cell.value) for cell in next(sheet.iter_rows(min_row=1, max_row=1))]
    rows: list[dict[str, Any]] = []
    for row in sheet.iter_rows(min_row=2, values_only=True):
        item = {headers[i]: row[i] if i < len(row) else None for i in range(len(headers))}
        if any(clean(value) for value in item.values()):
            rows.append(item)
    return rows


def orgs_from_row(row: dict[str, Any]) -> list[dict[str, str]]:
    orgs: list[dict[str, str]] = []
    for idx in range(1, 6):
        org = {
            "type": clean(row.get(f"org_{idx}_type")) or "조직",
            "name": clean(row.get(f"org_{idx}_name")),
            "owner": clean(row.get(f"org_{idx}_owner")),
            "role": clean(row.get(f"org_{idx}_role")),
        }
        if org["name"] or org["owner"]:
            orgs.append(org)
    return orgs


def convert(workbook_path: Path, output_path: Path) -> dict[str, Any]:
    wb = load_workbook(workbook_path, data_only=True)
    sheet = wb["01_Core_Tech_Data"] if "01_Core_Tech_Data" in wb.sheetnames else wb[wb.sheetnames[0]]
    rows = table_rows(sheet)

    output_rows: list[dict[str, Any]] = []
    for idx, row in enumerate(rows, start=1):
        if clean(row.get("active_yn")).upper() == "N":
            continue
        l1 = clean(row.get("level1"))
        l2 = clean(row.get("level2"))
        l3 = clean(row.get("level3"))
        if not all([l1, l2, l3]):
            continue
        output_rows.append(
            {
                "id": clean(row.get("tech_id")) or f"T-{idx:04d}",
                "l1": l1,
                "l2": l2,
                "l3": l3,
                "sticker": "",
                "summary": clean(row.get("기술요약") or row.get("summary")),
                "tags": clean(row.get("핵심기술_태그") or row.get("기술스티커") or row.get("대표기술명")),
                "isCore": as_bool(row.get("is_core")),
                "growthStage": norm_optional(row.get("growth_stage")),
                "researchStage": norm_optional(row.get("research_stage")),
                "orgs": orgs_from_row(row),
            }
        )

    payload = {
        "schemaVersion": 3,
        "source": str(workbook_path),
        "updatedAt": datetime.now(timezone.utc).isoformat(),
        "orgMaster": {"groups": DEFAULT_RESEARCH_GROUPS, "teams": DEFAULT_TEAMS},
        "rows": output_rows,
    }
    output_path.parent.mkdir(parents=True, exist_ok=True)
    output_path.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
    return payload


def main() -> None:
    parser = argparse.ArgumentParser(description="Convert Core Tech Excel template to HTML JSON.")
    parser.add_argument("workbook", type=Path, help="Path to PTRS Core Tech Excel workbook")
    parser.add_argument("--output", type=Path, default=DEFAULT_OUTPUT, help=f"Output JSON path. Default: {DEFAULT_OUTPUT}")
    args = parser.parse_args()
    payload = convert(args.workbook, args.output)
    print(f"saved: {args.output}")
    print(f"rows: {len(payload['rows'])}")


if __name__ == "__main__":
    main()
